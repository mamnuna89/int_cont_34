from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.urls import reverse

from .models import Risk, ControlPoint, Department, Division, ProcessDiagram
from .forms import RiskForm, ControlPointForm

import openpyxl

# 👉 Главная страница модуля внутреннего контроля
def control_index(request):
    return render(request, 'control_app/control_index.html')

def process_list(request):
    processes = ProcessDiagram.objects.select_related('department', 'division').all()
    return render(request, 'control_app/process_list.html', {'processes': processes})

def process_create(request):
    if request.method == 'POST':
        form = ProcessDiagramForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('control:process_list')
    else:
        form = ProcessDiagramForm()
    return render(request, 'control_app/process_form.html', {'form': form})

def process_edit(request, pk):
    process = get_object_or_404(ProcessDiagram, pk=pk)
    if request.method == 'POST':
        form = ProcessDiagramForm(request.POST, instance=process)
        if form.is_valid():
            form.save()
            return redirect('control:process_list')
    else:
        form = ProcessDiagramForm(instance=process)
    return render(request, 'control_app/process_form.html', {'form': form})

def process_delete(request, pk):
    process = get_object_or_404(ProcessDiagram, pk=pk)
    process.delete()
    return redirect('control:process_list')

def export_processes_excel(request):
    processes = ProcessDiagram.objects.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Code', 'Name', 'Department', 'Division'])
    for p in processes:
        ws.append([p.code, p.name, p.department.name if p.department else '', p.division.name if p.division else ''])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=processes.xlsx'
    wb.save(response)
    return response

# 👉 Отображение списка рисков
def risk_list(request):
    risks = Risk.objects.all()
    department_filter = request.GET.get('department')
    risk_type_filter = request.GET.get('risk_type')
    level_filter = request.GET.get('level')

    if department_filter:
        risks = risks.filter(department__icontains=department_filter)
    if risk_type_filter:
        risks = risks.filter(risk_type__icontains=risk_type_filter)
    if level_filter:
        try:
            level_value = int(level_filter)
            risks = risks.filter(level=level_value)
        except ValueError:
            pass

    for risk in risks:
        if risk.level is not None:
            if risk.level <= 6:
                risk.color = 'green'
            elif 7 <= risk.level <= 14:
                risk.color = 'yellow'
            else:
                risk.color = 'red'
        else:
            risk.color = 'gray'

    return render(request, 'control_app/control_risk_list.html', {
        'risks': risks,
        
    })

# 👉 Создание нового риска
def risk_create(request):
    if request.method == 'POST':
        form = RiskForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('control_risk_list')
    else:
        form = RiskForm()
    return render(request, 'control_app/control_risk_form.html', {
        'form': form,
        
    })

# 👉 Редактирование риска
def risk_edit(request, risk_id):
    risk = get_object_or_404(Risk, id=risk_id)
    if request.method == 'POST':
        form = RiskForm(request.POST, instance=risk)
        if form.is_valid():
            form.save()
            return redirect('control_risk_list')
    else:
        form = RiskForm(instance=risk)
    return render(request, 'control_app/control_risk_form.html', {
        'form': form,
        'risk': risk,
        
    })

# 👉 Удаление риска
@require_POST
def risk_delete(request, pk):
    risk = get_object_or_404(Risk, pk=pk)
    risk.delete()
    return redirect('control_risk_list')

# 👉 Экспорт рисков в Excel
def export_risks_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Risks"
    ws.append(['№', 'Code', 'Name', 'Type', 'Source', 'Registration Date', 'Department',
               'Owner', 'Process', 'Probability', 'Impact', 'Level'])

    risks = Risk.objects.all()
    for i, risk in enumerate(risks, start=1):
        ws.append([
            i,
            risk.risk_code,
            risk.name,
            risk.risk_type,
            risk.source,
            risk.registered_at.strftime('%Y-%m-%d') if risk.registered_at else '',
            str(risk.department),
            risk.owner,
            risk.process,
            risk.probability,
            risk.impact,
            risk.level
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=internal_control_risks.xlsx'
    wb.save(response)
    return response

# 👉 Структура департаментов
def department_structure(request):
    departments = Department.objects.prefetch_related(
        'divisions__processdiagram_set'
    ).all()
    
    return render(request, 'control_app/department_structure.html', {
        'departments': departments,
    
    })

# 👉 Обзор карты процессов
def process_map_overview(request):
    departments = Department.objects.prefetch_related('divisions').all()
    return render(request, 'control_app/process_map_overview.html', {
        'departments': departments,
        
    })

# 👉 Список всех диаграмм
def diagram_list(request):
    diagrams = ProcessDiagram.objects.select_related('department', 'division').all()
    return render(request, 'control_app/process_list.html', {
        'diagrams': diagrams,
        
    })


# 👉 Открытие редактора для новой схемы
def editor_view(request):
    departments = Department.objects.prefetch_related('divisions').all()
    return render(request, 'control_app/bpmn_editor.html', {
        'departments': departments,
        
    })

# 👉 Сохранение схемы
@require_POST
def save_process_diagram(request):
    name = request.POST.get('name')
    department_id = request.POST.get('department_id')
    division_id = request.POST.get('division_id')
    bpmn_xml = request.POST.get('bpmn_xml')

    department = Department.objects.get(id=department_id)
    division = Division.objects.get(id=division_id)

    diagram = ProcessDiagram.objects.create(
        name=name,
        department=department,
        division=division,
        bpmn_xml=bpmn_xml,
        created_by=request.user if request.user.is_authenticated else None
    )
    return JsonResponse({'status': 'success', 'diagram_id': diagram.id})


# 👉 Просмотр схемы
def diagram_view(request, diagram_id):
    diagram = get_object_or_404(ProcessDiagram, id=diagram_id)
    return render(request, 'control_app/diagram_view.html', {
        'diagram': diagram,
        'bpmn_xml': diagram.bpmn_xml
    })

# 👉 Редактирование схемы
def edit_diagram(request, diagram_id):
    diagram = get_object_or_404(ProcessDiagram, id=diagram_id)

    if request.method == 'POST':
        # ❗ Не изменяем department и division — чтобы code не сбивался
        diagram.name = request.POST.get('name')
        diagram.bpmn_xml = request.POST.get('bpmn_xml')
        diagram.save()
        return JsonResponse({'status': 'success'})

    departments = Department.objects.prefetch_related('divisions').all()
    return render(request, 'control_app/bpmn_editor.html', {
        'departments': departments,
        'diagram': diagram,
        
    })



# 👉 Удаление схемы
@require_POST
def delete_diagram(request, diagram_id):
    diagram = get_object_or_404(ProcessDiagram, id=diagram_id)
    diagram.delete()
    return HttpResponseRedirect(reverse('diagram_list'))

# 👉 Экспорт всех рисков

def control_export_risks_excel(request):
    risks = Risk.objects.all()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Internal Control Risks"

    headers = [
        '№', 'Code', 'Name', 'Risk Type', 'Source',
        'Registration Date', 'Department', 'Owner',
        'Process', 'Probability', 'Impact', 'Risk Level'
    ]
    sheet.append(headers)

    for idx, risk in enumerate(risks, start=1):
        sheet.append([
            idx,
            risk.risk_code,
            risk.name,
            risk.risk_type,
            risk.source,
            risk.registered_at.strftime('%Y-%m-%d') if risk.registered_at else '',
            str(risk.department),
            risk.owner,
            risk.process,
            risk.probability,
            risk.impact,
            risk.level
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=internal_control_risks.xlsx'
    workbook.save(response)
    return response

# 👉 Список контрольных точек
def control_point_list(request):
    selected_department = request.GET.get('department')
    selected_process = request.GET.get('process')

    control_points = ControlPoint.objects.all()
    if selected_department:
        control_points = control_points.filter(division__department__name=selected_department)
    if selected_process:
        control_points = control_points.filter(process__icontains=selected_process)

    departments = Department.objects.prefetch_related('divisions__processdiagram_set')


    return render(request, 'control_app/control_point_list.html', {
        'control_points': control_points,
        'departments': departments,
        'selected_department': selected_department,
        'selected_process': selected_process,
        
    })

# 👉 Добавление точки контроля
def control_point_create(request):
    if request.method == 'POST':
        form = ControlPointForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('control_point_list')
    else:
        form = ControlPointForm()
    return render(request, 'control_app/control_point_form.html', {
        'form': form,
        
    })

# 👉 Редактирование точки контроля
def control_point_edit(request, pk):
    point = get_object_or_404(ControlPoint, pk=pk)
    if request.method == 'POST':
        form = ControlPointForm(request.POST, instance=point)
        if form.is_valid():
            form.save()
            return redirect('control_point_list')
    else:
        form = ControlPointForm(instance=point)
    return render(request, 'control_app/control_point_form.html', {
        'form': form,
        
    })

# 👉 Удаление точки контроля
@require_POST
def control_point_delete(request, pk):
    point = get_object_or_404(ControlPoint, pk=pk)
    point.delete()
    return redirect('control_point_list')

# 👉 Получение рисков по процессу
def get_risks_by_process(request):
    process = request.GET.get('process')
    risks = Risk.objects.filter(process__iexact=process)
    data = [{'id': r.id, 'name': f"{r.risk_code} — {r.name}"} for r in risks]
    return JsonResponse({'risks': data})

# 👉 Экспорт точек контроля
def export_control_points_excel(request):
    control_points = ControlPoint.objects.all()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Control Points"

    headers = [
        'Process', 'Related Risk', 'Control Action', 'Control Procedure',
        'Control Type', 'Frequency', 'Responsible Person',
        'Control Method', 'Implemented'
    ]
    sheet.append(headers)

    for p in control_points:
        sheet.append([
            p.process,
            f"{p.related_risk.risk_code} — {p.related_risk.name}" if p.related_risk else "",
            p.control_action,
            p.control_procedure,
            p.get_control_type_display(),
            p.frequency,
            p.responsible_person,
            p.get_control_method_display(),
            'Yes' if p.implemented else 'No'
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=control_points.xlsx'
    workbook.save(response)
    return response
